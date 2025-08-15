C"""
Excel Data Handling with openpyxl
---------------------------------
Project Overview:
This script demonstrates how to read, write, and search data in an Excel file
using Python's `openpyxl` library.

Key Features:
1. Open and read Excel data.
2. Modify specific cell values.
3. Save changes back to the file.
4. Search for a specific test case row ("TEST CASE 2").
5. Extract that row’s data into a Python dictionary.

Tools & Skills:
- Python Programming
- openpyxl Library
- Excel File Handling
- Data Extraction & Manipulation
"""

import openpyxl

# -------------------- Step 1: Load the Excel file --------------------
file_path = r"C:\Users\Comp10\PycharmProjects\PythonProject\exceldemo.xlsx"
book = openpyxl.load_workbook(file_path)
sheet = book.active  # Get active sheet
data_dict = {}       # Dictionary to store extracted row data

# -------------------- Step 2: Read a specific cell --------------------
cell = sheet.cell(row=1, column=2)
print("Value in row 1, column 2:", cell.value)

# -------------------- Step 3: Modify a cell value --------------------
sheet.cell(row=2, column=2).value = "Rahul"
print("Updated value in row 2, column 2:", sheet.cell(row=2, column=2).value)

# Save the updated Excel file
book.save("exceldemo.xlsx")

# -------------------- Step 4: Display sheet dimensions --------------------
print("Total Rows:", sheet.max_row)
print("Total Columns:", sheet.max_column)

# -------------------- Step 5: Access a cell using Excel notation --------------------
print("Value in A5:", sheet['A5'].value)

# -------------------- Step 6: Search for 'TEST CASE 2' and extract row data --------------------
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TEST CASE 2":
        for j in range(2, sheet.max_column + 1):
            header = sheet.cell(row=1, column=j).value  # Column name
            value = sheet.cell(row=i, column=j).value   # Corresponding value
            data_dict[header] = value

# -------------------- Step 7: Print extracted data --------------------
print("Extracted Data Dictionary:", data_dict)
